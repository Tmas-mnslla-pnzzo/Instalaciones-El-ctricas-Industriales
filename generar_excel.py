"""
generar_excel.py
Usa tus módulos existentes (calculos_electricos.py, selector.py)
para leer datos_proyecto.json + datos_iluminacion.json
y generar el Excel de cálculos automáticamente.

Ejecutar: python generar_excel.py
Genera:   calculo_electrico_ponzoni.xlsx
          informe_errores.txt
"""
import json, math, traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from calculos_electricos import CargasElectricas, FactorPotencia, Conductores
from selector import SelectorComercial

# ═══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

class Logger:
    """
    Registra cada paso del cálculo.
    Niveles: INFO, WARN, ERROR
    Al final vuelca todo a informe_errores.txt y lo imprime resumido.
    """
    NIVELES = {"INFO": 0, "WARN": 1, "ERROR": 2}

    def __init__(self):
        self.entradas = []          # lista de (nivel, contexto, mensaje)
        self.contexto_actual = ""   # ej: "CCM1 > Motor B1 > conductor"

    def contexto(self, texto):
        self.contexto_actual = texto

    def info(self, msg):
        self.entradas.append(("INFO",  self.contexto_actual, msg))

    def warn(self, msg):
        self.entradas.append(("WARN",  self.contexto_actual, msg))
        print(f"  ⚠  [{self.contexto_actual}] {msg}")

    def error(self, msg):
        self.entradas.append(("ERROR", self.contexto_actual, msg))
        print(f"  ✖  [{self.contexto_actual}] {msg}")

    # ── Validadores reutilizables ────────────────────────────────────────────

    def validar_numero(self, valor, nombre, contexto="", minimo=None, maximo=None):
        """Devuelve True si el valor es un número válido y dentro de rango."""
        ctx = contexto or self.contexto_actual
        if valor is None:
            self.entradas.append(("ERROR", ctx, f"{nombre} es None → celda vacía"))
            print(f"  ✖  [{ctx}] {nombre} es None → celda vacía en el Excel")
            return False
        if not isinstance(valor, (int, float)):
            self.entradas.append(("ERROR", ctx, f"{nombre} no es número: {repr(valor)}"))
            print(f"  ✖  [{ctx}] {nombre} no es número: {repr(valor)}")
            return False
        if math.isnan(valor) or math.isinf(valor):
            self.entradas.append(("ERROR", ctx, f"{nombre} = {valor} (NaN o Inf)"))
            print(f"  ✖  [{ctx}] {nombre} = {valor} (NaN o Inf) → celda vacía en el Excel")
            return False
        if minimo is not None and valor < minimo:
            self.entradas.append(("WARN", ctx, f"{nombre} = {valor} < mínimo esperado {minimo}"))
            print(f"  ⚠  [{ctx}] {nombre} = {valor} < mínimo esperado {minimo}")
        if maximo is not None and valor > maximo:
            self.entradas.append(("WARN", ctx, f"{nombre} = {valor} > máximo esperado {maximo}"))
            print(f"  ⚠  [{ctx}] {nombre} = {valor} > máximo esperado {maximo}")
        return True

    def validar_resultado_conductor(self, res, contexto):
        """Valida el dict que devuelve Conductores.dimensionamiento_iterativo()."""
        if not res.get("cumple_todas_las_condiciones", False):
            self.entradas.append(("ERROR", contexto,
                f"dimensionamiento_iterativo no encontró sección válida: {res.get('detalle','')}"))
            print(f"  ✖  [{contexto}] No se encontró sección de cable válida → {res.get('detalle','')}")
            return False
        if res.get("seccion_mm2") is None:
            self.entradas.append(("ERROR", contexto,
                "seccion_mm2 = None (se usará fallback 240 mm²)"))
            print(f"  ✖  [{contexto}] seccion_mm2 = None → se usa fallback 240 mm²")
            return False
        return True

    # ── Exportar informe ─────────────────────────────────────────────────────

    def exportar(self, ruta="informe_errores.txt"):
        errores  = [e for e in self.entradas if e[0] == "ERROR"]
        warnings = [e for e in self.entradas if e[0] == "WARN"]
        infos    = [e for e in self.entradas if e[0] == "INFO"]

        with open(ruta, "w", encoding="utf-8") as f:
            f.write("=" * 70 + "\n")
            f.write(f"  INFORME DE EJECUCIÓN — generar_excel.py\n")
            f.write(f"  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write("=" * 70 + "\n\n")

            f.write(f"RESUMEN: {len(errores)} errores | {len(warnings)} advertencias | {len(infos)} pasos OK\n\n")

            if errores:
                f.write("─" * 70 + "\n")
                f.write(f"ERRORES ({len(errores)}) — estas celdas estarán vacías o con valor incorrecto\n")
                f.write("─" * 70 + "\n")
                for _, ctx, msg in errores:
                    f.write(f"  [ERROR] {ctx}\n         {msg}\n\n")

            if warnings:
                f.write("─" * 70 + "\n")
                f.write(f"ADVERTENCIAS ({len(warnings)}) — revisar valores\n")
                f.write("─" * 70 + "\n")
                for _, ctx, msg in warnings:
                    f.write(f"  [WARN]  {ctx}\n         {msg}\n\n")

            f.write("─" * 70 + "\n")
            f.write(f"PASOS COMPLETADOS ({len(infos)})\n")
            f.write("─" * 70 + "\n")
            for _, ctx, msg in infos:
                f.write(f"  [OK]    {ctx} — {msg}\n")

        print(f"\n{'='*50}")
        print(f"  Informe guardado en: {ruta}")
        print(f"  Errores:       {len(errores)}")
        print(f"  Advertencias:  {len(warnings)}")
        print(f"  Pasos OK:      {len(infos)}")
        print(f"{'='*50}\n")
        return len(errores)


# Instancia global
LOG = Logger()


# ═══════════════════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════

LOG.contexto("Carga de archivos JSON")
try:
    with open("datos_proyecto.json", encoding="utf-8") as f:
        PROYECTO = json.load(f)
    LOG.info("datos_proyecto.json cargado OK")
except Exception as e:
    LOG.error(f"No se pudo leer datos_proyecto.json: {e}")
    raise SystemExit("Archivo de proyecto no encontrado. Abortando.")

try:
    with open("datos_iluminacion.json", encoding="utf-8") as f:
        ILUMINACION = json.load(f)["circuitos"]
    LOG.info("datos_iluminacion.json cargado OK")
except Exception as e:
    LOG.error(f"No se pudo leer datos_iluminacion.json: {e}")
    ILUMINACION = []

try:
    with open("valores_comerciales.json", encoding="utf-8") as f:
        COMERCIALES = json.load(f)
    LOG.info("valores_comerciales.json cargado OK")
except Exception as e:
    LOG.error(f"No se pudo leer valores_comerciales.json: {e}")
    raise SystemExit("Catálogo comercial no encontrado. Abortando.")

INFO        = PROYECTO["informacion_general"]
CCMS        = PROYECTO["ccms"]
CONDUCTORES = PROYECTO["conductores"]
SELECTOR    = SelectorComercial("valores_comerciales.json")


# ═══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════════════════════

def borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

C = {
    "titulo":    "1F3864",
    "encabezado":"2E75B6",
    "celeste":   "D6E4F0",
    "gris":      "D9D9D9",
    "verde":     "E2EFDA",
    "amarillo":  "FFF2CC",
    "rojo":      "FFDCE0",   # ← celdas con error
}

def h1(ws, row, col, texto, span=1):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=texto)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=C["titulo"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = borde()

def h2(ws, row, col, texto):
    c = ws.cell(row=row, column=col, value=texto)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c.fill      = PatternFill("solid", fgColor=C["encabezado"])
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = borde()

def cel(ws, row, col, valor, fmt=None, bg=None, bold=False, izq=False):
    """Escribe una celda. Si el valor es None o NaN lo marca en rojo."""
    es_invalido = (
        valor is None
        or (isinstance(valor, float) and (math.isnan(valor) or math.isinf(valor)))
    )
    if es_invalido:
        valor = "⚠ ERROR"
        bg    = C["rojo"]

    c = ws.cell(row=row, column=col, value=valor)
    c.font      = Font(name="Arial", size=9, bold=bold,
                       color="CC0000" if es_invalido else "000000")
    c.alignment = Alignment(horizontal="left" if izq else "center", vertical="center")
    c.border    = borde()
    if fmt and not es_invalido:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


# ═══════════════════════════════════════════════════════════════════════════════
# CÁLCULOS
# ═══════════════════════════════════════════════════════════════════════════════

def calcular_ccms():
    resultados = []
    V      = INFO["tension_red_V"]
    fp_obj = INFO["cos_phi_objetivo"]

    for ccm in CCMS:
        LOG.contexto(f"CCM {ccm['id']}")
        fs = ccm["factor_simultaneidad"]
        motores_calc = []
        P_neto = Q_neto = 0.0

        for m in ccm["motores"]:
            ctx = f"CCM {ccm['id']} > Motor {m['id']}"
            LOG.contexto(ctx)

            Pa  = m.get("potencia_activa_kW")
            Qr  = m.get("potencia_reactiva_kVAr")
            Sa  = m.get("potencia_aparente_kVA")
            Ib  = m.get("corriente_A")
            cp  = m.get("cos_phi")
            fu  = m.get("factor_utilizacion")

            # Validar cada campo crítico
            ok = all([
                LOG.validar_numero(Pa,  "potencia_activa_kW",     ctx, minimo=0),
                LOG.validar_numero(Qr,  "potencia_reactiva_kVAr", ctx, minimo=0),
                LOG.validar_numero(Sa,  "potencia_aparente_kVA",  ctx, minimo=0),
                LOG.validar_numero(Ib,  "corriente_A",            ctx, minimo=0, maximo=2000),
                LOG.validar_numero(cp,  "cos_phi",                ctx, minimo=0.5, maximo=1.0),
                LOG.validar_numero(fu,  "factor_utilizacion",     ctx, minimo=0.1, maximo=1.0),
            ])

            # Consistencia: S debe ≈ P/cos_phi
            if ok and abs(Sa - Pa/cp) > 0.5:
                LOG.warn(f"potencia_aparente_kVA={Sa:.3f} no coincide con P/cos_phi={Pa/cp:.3f}")

            motores_calc.append({**m, "Pa": Pa, "Qr": Qr, "Sa": Sa, "Ib": Ib})
            if Pa: P_neto += Pa
            if Qr: Q_neto += Qr

            if ok:
                LOG.info(f"Motor {m['id']}: Ib={Ib:.3f}A  Pa={Pa:.3f}kW  Sa={Sa:.3f}kVA")

        # Totales del CCM
        LOG.contexto(f"CCM {ccm['id']} > Totales")
        S_neto   = math.sqrt(P_neto**2 + Q_neto**2) if (P_neto or Q_neto) else 0
        cos_neto = P_neto / S_neto if S_neto else 0
        I_neto   = (S_neto * 1000) / (math.sqrt(3) * V) if V else 0

        P_cor = P_neto * fs
        Q_cor = Q_neto * fs
        S_cor = math.sqrt(P_cor**2 + Q_cor**2)
        cos_cor = P_cor / S_cor if S_cor else 0
        I_cor   = (S_cor * 1000) / (math.sqrt(3) * V) if V else 0

        try:
            Qc = max(0, FactorPotencia.banco_capacitores(P_cor, cos_neto, fp_obj))
            LOG.info(f"Banco capacitores: Qc={Qc:.4f} kVAr")
        except Exception as e:
            LOG.error(f"FactorPotencia.banco_capacitores falló: {e}")
            Qc = None

        LOG.validar_numero(I_cor,   "I_cor total CCM",  f"CCM {ccm['id']}", minimo=0, maximo=5000)
        LOG.validar_numero(cos_cor, "cos_phi corregido",f"CCM {ccm['id']}", minimo=0.5, maximo=1.0)
        LOG.info(f"Total neto: P={P_neto:.2f}kW  S={S_neto:.2f}kVA  I={I_neto:.2f}A")
        LOG.info(f"Total corregido (fs={fs}): P={P_cor:.2f}kW  S={S_cor:.2f}kVA  I={I_cor:.2f}A")

        resultados.append({
            "id": ccm["id"], "tablero_origen": ccm["tablero_origen"], "fs": fs,
            "motores": motores_calc,
            "P_neto": P_neto, "Q_neto": Q_neto, "S_neto": S_neto,
            "cos_neto": cos_neto, "I_neto": I_neto,
            "P_cor": P_cor, "Q_cor": Q_cor, "S_cor": S_cor,
            "cos_cor": cos_cor, "I_cor": I_cor,
            "Qc": Qc,
        })
    return resultados


def calcular_iluminacion():
    resultados = []
    for i, lu in enumerate(ILUMINACION):
        ctx = f"Iluminación > {lu.get('tablero','?')} ({lu.get('marca','?')})"
        LOG.contexto(ctx)

        P_W  = lu.get("potencia_W")
        cant = lu.get("cantidad")
        V_lu = lu.get("tension_V")
        cp   = lu.get("cos_phi")
        fu   = lu.get("factor_utilizacion")
        fs   = lu.get("factor_simultaneidad")

        ok = all([
            LOG.validar_numero(P_W,  "potencia_W",           ctx, minimo=0.1, maximo=5000),
            LOG.validar_numero(cant, "cantidad",              ctx, minimo=1),
            LOG.validar_numero(V_lu, "tension_V",            ctx, minimo=100, maximo=500),
            LOG.validar_numero(cp,   "cos_phi",              ctx, minimo=0.5, maximo=1.0),
            LOG.validar_numero(fu,   "factor_utilizacion",   ctx, minimo=0.1, maximo=1.0),
            LOG.validar_numero(fs,   "factor_simultaneidad", ctx, minimo=0.1, maximo=1.0),
        ])

        if ok:
            P_kW  = P_W * cant / 1000.0
            I     = (P_kW * 1000) / (V_lu * cp)
            P_cor = P_kW * fu * fs
            LOG.info(f"P_total={P_kW:.3f}kW  I={I:.3f}A  P_cor={P_cor:.3f}kW")
        else:
            P_kW = P_cor = I = None

        resultados.append({
            "tablero": lu.get("tablero"), "marca": lu.get("marca"),
            "V": V_lu, "cos_phi": cp,
            "cantidad": cant, "fu": fu, "fs": fs,
            "P_kW": P_kW, "I": I, "P_cor": P_cor,
        })
    return resultados


def calcular_conductores(ccms_res, ilu_res):
    corrientes = {}
    for ccm in ccms_res:
        corrientes[ccm["id"]] = ccm["I_cor"]
        for m in ccm["motores"]:
            corrientes[m["id"]] = m["Ib"]
    for lu in ilu_res:
        if lu["tablero"] and lu["I"]:
            corrientes[lu["tablero"]] = lu["I"]

    V      = INFO["tension_red_V"]
    icc_max= INFO["potencia_cc_mva"] * 1e6 / (math.sqrt(3) * V)
    t_cc   = INFO["tiempo_cc_s"]
    k      = INFO["k_aislante"]
    dU_max = INFO["delta_u_max_V"]

    resultados = []
    for cond in CONDUCTORES:
        ctx = f"Conductor {cond['origen']} → {cond['destino']}"
        LOG.contexto(ctx)

        Ib = corrientes.get(cond["destino"])
        if Ib is None:
            LOG.warn(f"No se encontró corriente para destino '{cond['destino']}' → se usa fallback 50 A")
            Ib = 50.0
        else:
            LOG.validar_numero(Ib, "Ib", ctx, minimo=0.1, maximo=5000)

        # Dimensionamiento iterativo
        try:
            res = Conductores.dimensionamiento_iterativo(
                ib=Ib,
                longitud_km=cond["longitud_m"] / 1000.0,
                cos_phi=0.85,
                icc_max=icc_max,
                tiempo_cc=t_cc,
                delta_u_max=dU_max,
                k_aislante=k,
                seccion_maxima=240.0,
            )
            LOG.validar_resultado_conductor(res, ctx)
            sec   = res["seccion_mm2"] or 240.0
            n_par = res["cantidad_de_cables_en_paralelo"]
            LOG.info(f"Sección: {sec}mm² x{n_par} | {res.get('detalle','')}")
        except Exception as e:
            LOG.error(f"dimensionamiento_iterativo lanzó excepción: {e}")
            sec, n_par = 240.0, 1

        # Sección comercial
        try:
            sec_com = SELECTOR.obtener_inmediato_superior("secciones_cables_mm2", sec)
        except Exception as e:
            LOG.error(f"SelectorComercial falló para sección {sec}mm²: {e}")
            sec_com = sec

        # Iz
        iz_base  = Conductores.CATALOGO_IZ_BASE.get(sec_com, 0)
        iz_total = iz_base * cond["fa"] * cond["ft"] * n_par
        if iz_base == 0:
            LOG.warn(f"Sección {sec_com}mm² no está en CATALOGO_IZ_BASE → Iz=0")

        # Caída de tensión
        try:
            rho    = 0.0225
            xc     = 0.08
            sin_phi= math.sin(math.acos(0.85))
            rc     = (rho * 1000 / sec_com) / n_par
            xc_eq  = xc / n_par
            dU     = math.sqrt(3) * Ib * (cond["longitud_m"]/1000) * (rc*0.85 + xc_eq*sin_phi)
            dU_pct = (dU / V) * 100
            if dU_pct > 5.0:
                LOG.warn(f"ΔU={dU_pct:.2f}% supera el 5% admisible (L={cond['longitud_m']}m, Ib={Ib:.1f}A, S={sec_com}mm²)")
            else:
                LOG.info(f"ΔU={dU_pct:.2f}% ✓  Iz={iz_total:.1f}A  Ib={Ib:.1f}A")
        except Exception as e:
            LOG.error(f"Cálculo de caída de tensión falló: {e}")
            dU = dU_pct = None

        # Termomagnética
        try:
            In_term = SELECTOR.obtener_inmediato_superior("termomagneticas_A", Ib * 1.25)
        except Exception as e:
            LOG.warn(f"No se encontró termomagnética para {Ib*1.25:.1f}A: {e}")
            In_term = None

        resultados.append({
            **cond,
            "Ib": Ib, "seccion": sec_com, "n_paralelo": n_par,
            "Iz": iz_total, "dU_V": dU, "dU_pct": dU_pct,
            "In_term": In_term,
            "cumple": (dU_pct is not None and dU_pct <= 5.0),
        })
    return resultados


def calcular_totales(ccms_res, ilu_res):
    LOG.contexto("Totales generales")
    P_mot = sum(c["P_neto"] for c in ccms_res if c["P_neto"])
    Q_mot = sum(c["Q_neto"] for c in ccms_res if c["Q_neto"])
    P_ilu = sum(l["P_kW"]   for l in ilu_res   if l["P_kW"])
    P_tot = P_mot + P_ilu
    Q_tot = Q_mot
    S_tot = math.sqrt(P_tot**2 + Q_tot**2) if (P_tot or Q_tot) else 0
    cos_t = P_tot / S_tot if S_tot else 0
    S_res = S_tot * INFO["factor_reserva"]

    LOG.validar_numero(S_tot, "S total", minimo=1)
    LOG.validar_numero(cos_t, "cos_phi global", minimo=0.5, maximo=1.0)

    try:
        trafo = SELECTOR.obtener_inmediato_superior("transformadores_kVA", S_res)
        LOG.info(f"Transformador seleccionado: {trafo} kVA (S_res={S_res:.1f} kVA)")
    except Exception as e:
        LOG.error(f"No se encontró transformador para {S_res:.1f} kVA: {e}")
        trafo = None

    LOG.info(f"P_tot={P_tot:.2f}kW  S_tot={S_tot:.2f}kVA  cos_phi={cos_t:.4f}  S_res={S_res:.2f}kVA")
    return {
        "P_mot": P_mot, "Q_mot": Q_mot, "P_ilu": P_ilu,
        "P_tot": P_tot, "Q_tot": Q_tot, "S_tot": S_tot,
        "cos_t": cos_t, "S_res": S_res, "trafo": trafo,
        "cant_luminarias": sum(l["cantidad"] for l in ILUMINACION if l.get("cantidad")),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ESCRITURA EXCEL  (igual que antes, la función cel() ya marca errores en rojo)
# ═══════════════════════════════════════════════════════════════════════════════

def escribir_excel(ccms_res, ilu_res, cond_res, tot):
    wb = Workbook()

    # ── Hoja 1: Calculos ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Calculos"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([14,12,8,10,10,12,12,12,8,10,10,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    h1(ws, r, 1, f"CÁLCULOS ELÉCTRICOS — {INFO['nombre_proyecto']}", span=12)
    ws.row_dimensions[r].height = 22; r += 1
    cel(ws, r, 1, f"Autor: {INFO['autor']}  |  Tensión: {INFO['tension_red_V']} V  |  fp obj: {INFO['cos_phi_objetivo']}", bg=C["gris"])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(r,1).alignment = Alignment(horizontal="left"); r += 2

    # Cargas motrices
    h1(ws, r, 1, "CARGAS MOTRICES", span=12); r += 1
    for j, txt in enumerate(["Tablero","Motor","Tensión [V]","Corriente [A]","cos φ",
                              "P Activa [kW]","P Reactiva [kVAr]","P Aparente [kVA]",
                              "Zcc [Ω]","Dist CC [km]","F. Utiliz.","F. Simult."], 1):
        h2(ws, r, j, txt)
    ws.row_dimensions[r].height = 28; r += 1

    for ccm in ccms_res:
        first = True
        for m in ccm["motores"]:
            bg_ccm = C["celeste"] if first else None
            cel(ws, r, 1,  ccm["id"] if first else "",   bg=bg_ccm, bold=first)
            cel(ws, r, 2,  m["id"])
            cel(ws, r, 3,  INFO["tension_red_V"],         fmt="0.0")
            cel(ws, r, 4,  m["Ib"],                       fmt="0.000")
            cel(ws, r, 5,  m.get("cos_phi"),              fmt="0.000")
            cel(ws, r, 6,  m["Pa"],                       fmt="0.000")
            cel(ws, r, 7,  m["Qr"],                       fmt="0.000")
            cel(ws, r, 8,  m["Sa"],                       fmt="0.000")
            cel(ws, r, 9,  m.get("zcc"),                  fmt="0.00")
            cel(ws, r, 10, m.get("distancia_cc_km"),      fmt="0.00")
            cel(ws, r, 11, m.get("factor_utilizacion"),   fmt="0.000")
            cel(ws, r, 12, ccm["fs"] if first else "",    fmt="0.000" if first else None, bg=bg_ccm)
            first = False; r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cel(ws, r, 1, f"Total neto {ccm['id']}:",   bg=C["amarillo"], bold=True, izq=True)
        cel(ws, r, 4, ccm["I_neto"],   fmt="0.000", bg=C["amarillo"])
        cel(ws, r, 5, ccm["cos_neto"], fmt="0.000", bg=C["amarillo"])
        cel(ws, r, 6, ccm["P_neto"],   fmt="0.000", bg=C["amarillo"])
        cel(ws, r, 7, ccm["Q_neto"],   fmt="0.000", bg=C["amarillo"])
        cel(ws, r, 8, ccm["S_neto"],   fmt="0.000", bg=C["amarillo"])
        for j in [3,9,10,11,12]: cel(ws, r, j, "", bg=C["amarillo"]); r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cel(ws, r, 1, f"Total corregido {ccm['id']}:", bg=C["verde"], bold=True, izq=True)
        cel(ws, r, 4, ccm["I_cor"],   fmt="0.000", bg=C["verde"])
        cel(ws, r, 5, ccm["cos_cor"], fmt="0.000", bg=C["verde"])
        cel(ws, r, 6, ccm["P_cor"],   fmt="0.000", bg=C["verde"])
        cel(ws, r, 7, ccm["Q_cor"],   fmt="0.000", bg=C["verde"])
        cel(ws, r, 8, ccm["S_cor"],   fmt="0.000", bg=C["verde"])
        for j in [3,9,10,11,12]: cel(ws, r, j, "", bg=C["verde"]); r += 1

    r += 1

    # Iluminación
    h1(ws, r, 1, "ILUMINACIÓN", span=12); r += 1
    for j, txt in enumerate(["Tablero","Luminaria","Tensión [V]","Corriente [A]","cos φ",
                              "P Activa [kW]","P Reactiva [kVAr]","P Aparente [kVA]",
                              "Cantidad","F. Utiliz.","F. Simult.",""], 1):
        h2(ws, r, j, txt)
    ws.row_dimensions[r].height = 25; r += 1

    for lu in ilu_res:
        cel(ws, r, 1,  lu["tablero"])
        cel(ws, r, 2,  lu["marca"])
        cel(ws, r, 3,  lu["V"],       fmt="0.0")
        cel(ws, r, 4,  lu["I"],       fmt="0.000")
        cel(ws, r, 5,  lu["cos_phi"], fmt="0.0")
        cel(ws, r, 6,  lu["P_kW"],    fmt="0.000")
        cel(ws, r, 7,  0.0,           fmt="0.0")
        cel(ws, r, 8,  lu["P_kW"],    fmt="0.000")
        cel(ws, r, 9,  lu["cantidad"])
        cel(ws, r, 10, lu["fu"],      fmt="0.0")
        cel(ws, r, 11, lu["fs"],      fmt="0.0")
        cel(ws, r, 12, ""); r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cel(ws, r, 1, "Total neto iluminación:", bg=C["amarillo"], bold=True, izq=True)
    cel(ws, r, 6, sum(l["P_kW"] or 0 for l in ilu_res), fmt="0.000", bg=C["amarillo"])
    cel(ws, r, 9, sum(l["cantidad"] or 0 for l in ilu_res), bg=C["amarillo"])
    for j in [3,4,5,7,8,10,11,12]: cel(ws, r, j, "", bg=C["amarillo"]); r += 2

    # Totales
    h1(ws, r, 1, "RESUMEN DE POTENCIAS Y TRANSFORMADOR", span=12); r += 1
    for j, txt in enumerate(["Concepto","cos φ","P Activa [kW]","P Reactiva [kVAr]","P Aparente [kVA]"], 1):
        h2(ws, r, j, txt); r += 1
    for fila in [
        ("TOTAL NETO [kVA]:",             tot["cos_t"], tot["P_tot"], tot["Q_tot"], tot["S_tot"]),
        ("TOTAL CON RESERVA (30%) [kVA]:", None, None, None, tot["S_res"]),
        ("TRANSFORMADOR COMERCIAL [kVA]:", None, None, None, tot["trafo"]),
    ]:
        cel(ws, r, 1, fila[0], bg=C["celeste"], bold=True, izq=True)
        for j, v in enumerate(fila[1:], 2):
            cel(ws, r, j, v if v is not None else "-",
                fmt="0.000" if isinstance(v, float) else None, bg=C["celeste"])
        r += 1
    r += 2

    # Capacitores
    h1(ws, r, 1, "BANCO DE CAPACITORES POR CCM", span=8); r += 1
    for j, txt in enumerate(["CCM","Corrección FP [kVAr]","Pot. CC [MVA]",
                              "In [A]","I Prot [A]","I Cresta [A]","K Conexión",""], 1):
        h2(ws, r, j, txt); r += 1
    for ccm in ccms_res:
        Qc  = ccm["Qc"]
        In  = (Qc * 1000 / (math.sqrt(3) * INFO["tension_red_V"])) if Qc else 0
        cel(ws, r, 1, ccm["id"], bold=True)
        cel(ws, r, 2, round(Qc, 4) if Qc else None, fmt="0.0000")
        cel(ws, r, 3, INFO["potencia_cc_mva"])
        cel(ws, r, 4, round(In, 3),          fmt="0.000")
        cel(ws, r, 5, round(In*1.43, 3),     fmt="0.000")
        cel(ws, r, 6, round(In*math.sqrt(2)*1.41, 3), fmt="0.000")
        cel(ws, r, 7, 3); cel(ws, r, 8, ""); r += 1
    r += 2

    # Conductores
    h1(ws, r, 1, "CONDUCTORES — DIMENSIONAMIENTO", span=12); r += 1
    for j, txt in enumerate(["Origen","Destino","Tipo","Sección [mm²]","N° Par.",
                              "Iz total [A]","Ib [A]","ΔU [V]","ΔU [%]",
                              "In term. [A]","Long. [m]","OK"], 1):
        h2(ws, r, j, txt)
    ws.row_dimensions[r].height = 25; r += 1

    for cond in cond_res:
        bg = C["verde"] if cond["cumple"] else C["amarillo"]
        cel(ws, r, 1,  cond["origen"])
        cel(ws, r, 2,  cond["destino"])
        cel(ws, r, 3,  cond["tipo"])
        cel(ws, r, 4,  cond["seccion"],           fmt="0.0",   bg=bg)
        cel(ws, r, 5,  cond["n_paralelo"])
        cel(ws, r, 6,  round(cond["Iz"],   1) if cond["Iz"]   else None, fmt="0.0")
        cel(ws, r, 7,  round(cond["Ib"],   2) if cond["Ib"]   else None, fmt="0.00")
        cel(ws, r, 8,  round(cond["dU_V"], 3) if cond["dU_V"] else None, fmt="0.000", bg=bg)
        cel(ws, r, 9,  round(cond["dU_pct"],2) if cond["dU_pct"] else None, fmt="0.00", bg=bg)
        cel(ws, r, 10, cond["In_term"])
        cel(ws, r, 11, cond["longitud_m"])
        cel(ws, r, 12, "✓" if cond["cumple"] else "⚠", bg=bg); r += 1

    # ── Hoja 2: Comerciales ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Comerciales")
    ws2.sheet_view.showGridLines = False
    h1(ws2, 1, 1, "CATÁLOGOS COMERCIALES", span=8)
    for j, txt in enumerate(["Transf. [kVA]","Condensadores","Conductores [A]",
                              "Conductores [mm²]","R [Ω/km]","X [Ω/km]",
                              "Termomagnéticas [A]","Fusibles [A]"], 1):
        h2(ws2, 2, j, txt)
        ws2.column_dimensions[get_column_letter(j)].width = 16
    datos_com = [
        (9,   "",  9.6,   1.0,  18.10,0.13,  6,  2),(15,"",13.0, 1.5,12.10,0.12,10,  4),
        (30,  "",  18.0,  2.5,   7.28,0.11, 16, 10),(45,"",24.0, 4.0, 4.55,0.10,20, 16),
        (75,  "",  31.0,  6.0,   3.03,0.10, 25, 20),(113,"",43.0,10.0,1.82,0.09,32, 25),
        (150, "",  59.0, 16.0,   1.14,0.08, 40, 35),(225,"",77.0,25.0,0.73,0.08,50, 50),
        (300, "",  96.0, 35.0,   0.52,0.08, 63, 63),(500,"",116.0,50.0,0.36,0.08,80,""),
        (750, "", 148.0, 70.0,   0.26,0.07,100,""),(1000,"",180.0,95.0,0.19,0.07,125,""),
        (1500,"", 210.0,120.0,   0.15,0.07,160,""),(2000,"",245.0,150.0,0.12,0.07,200,""),
        (2500,"", 285.0,185.0,   0.10,0.07,250,""),(3750,"",345.0,240.0,0.075,0.06,315,""),
    ]
    for i, fila in enumerate(datos_com, 3):
        for j, v in enumerate(fila, 1):
            c = ws2.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center")
            c.border = borde()
            if i % 2 == 0: c.fill = PatternFill("solid", fgColor=C["gris"])

    # ── Hoja 3: CortoCircuito ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("CortoCircuito")
    ws3.sheet_view.showGridLines = False
    h1(ws3, 1, 1, "CÁLCULO DE CORTOCIRCUITO", span=6)
    for j, txt in enumerate(["Tablero","Zcc total [Ω]","Icc máx [kA]",
                              "Icc mín [kA]","Secc. mín CC [mm²]","Estado"], 1):
        h2(ws3, 2, j, txt)
        ws3.column_dimensions[get_column_letter(j)].width = 18
    V_red = INFO["tension_red_V"]; Pcc = INFO["potencia_cc_mva"]*1e6; Zb = V_red**2/Pcc
    for i, nodo in enumerate(["T","TG","TS4","CCM1","CCM2","CCM3","CCM4"], 3):
        Zcc     = Zb*(i-2)*0.5
        Icc_max = V_red/(math.sqrt(3)*Zcc)/1000
        Icc_min = Icc_max*0.85
        sec_min = (Icc_max*1000*math.sqrt(INFO["tiempo_cc_s"]))/INFO["k_aislante"]
        ok      = sec_min < 240
        cel(ws3,i,1,nodo,bold=True); cel(ws3,i,2,round(Zcc,4),fmt="0.0000")
        cel(ws3,i,3,round(Icc_max,3),fmt="0.000"); cel(ws3,i,4,round(Icc_min,3),fmt="0.000")
        cel(ws3,i,5,round(sec_min,2),fmt="0.00")
        cel(ws3,i,6,"OK" if ok else "Verificar",bg=C["verde"] if ok else C["amarillo"])

    # ── Hoja 4: Resumen ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumen")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 38; ws4.column_dimensions["B"].width = 22
    h1(ws4, 1, 1, "RESUMEN EJECUTIVO", span=2)
    for i, (k, v) in enumerate([
        ("Proyecto",                   INFO["nombre_proyecto"]),
        ("Autor",                      INFO["autor"]),
        ("Tensión de red",             f"{INFO['tension_red_V']} V"),
        ("P activa total neta",        f"{tot['P_tot']:.2f} kW"   if tot['P_tot'] else "ERROR"),
        ("P aparente total neta",      f"{tot['S_tot']:.2f} kVA"  if tot['S_tot'] else "ERROR"),
        ("Factor de potencia global",  f"{tot['cos_t']:.4f}"      if tot['cos_t'] else "ERROR"),
        ("P con reserva 30%",          f"{tot['S_res']:.2f} kVA"  if tot['S_res'] else "ERROR"),
        ("Transformador seleccionado", f"{tot['trafo']} kVA"       if tot['trafo'] else "ERROR"),
        ("Total luminarias",           f"{tot['cant_luminarias']} unid."),
    ], 2):
        bg = C["celeste"] if i%2==0 else C["gris"]
        cel(ws4,i,1,k,bg=bg,bold=True,izq=True); cel(ws4,i,2,v,bg=bg)

    salida = "calculo_electrico_ponzoni.xlsx"
    wb.save(salida)
    print(f"✅  Excel generado: {salida}")
    return salida


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("⚙️  Calculando CCMs...")
    ccms_res = calcular_ccms()

    print("💡 Calculando iluminación...")
    ilu_res  = calcular_iluminacion()

    print("🔌 Dimensionando conductores...")
    cond_res = calcular_conductores(ccms_res, ilu_res)

    print("📊 Totalizando potencias...")
    tot      = calcular_totales(ccms_res, ilu_res)

    print(f"\n   P total neta  : {tot['P_tot']:.2f} kW")
    print(f"   S total neta  : {tot['S_tot']:.2f} kVA")
    print(f"   cos φ global  : {tot['cos_t']:.4f}")
    print(f"   Transformador : {tot['trafo']} kVA")

    escribir_excel(ccms_res, ilu_res, cond_res, tot)

    # Exportar informe — siempre, haya errores o no
    n_errores = LOG.exportar("informe_errores.txt")
    if n_errores > 0:
        print(f"\n⚠  Hay {n_errores} errores. Revisá informe_errores.txt para saber qué celdas están mal.")
    else:
        print("✅  Sin errores. Ver informe_errores.txt para el detalle completo.")